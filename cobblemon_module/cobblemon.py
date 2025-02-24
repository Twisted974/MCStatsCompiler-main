import json
import os
import pandas as pd
import numpy as np
import configparser
import openpyxl
import datetime
import paramiko
import math


def loadData(csvtoggle, csvpath, useftp, sftp, ftppath):
    df = pd.DataFrame()
    names = pd.read_csv('../data/names.csv')
    root_dirnames = []
    if useftp == "true":
        # Get directories
        root_dirnames = sftp.listdir(ftppath)
        
        for dirname in root_dirnames:
            if dirname[-1] == ".":
                continue
            # Go to the subfolder
            subfolder_path = f"{ftppath}/{dirname}"
            filenames = sftp.listdir(subfolder_path)
            
            for filename in filenames:
                if filename == "." or filename == "..":
                    continue
                print("Now processing", filename)
                
                # Download the file to process
                local_file = f"temp_{filename}"
                with open(local_file, "wb") as file:
                    sftp.get(f"{subfolder_path}/{filename}", local_file)
                
                with open(local_file, "r") as file:
                    data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                
                os.remove(local_file)
                
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[:]
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in names.csv, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                
                if not temp_df.empty:
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
                
            # No need to go back a directory in paramiko since `sftp` handles the current path.
    else:
        i = -1
        path = 'cobblemonplayerdata'
        for dirpath, dirnames, filenames in os.walk(path):
            if len(dirnames) > 0:
                root_dirnames = dirnames
            for filename in filenames:
                if filename == ".gitignore":
                    continue
                print("Now processing", filename)
                file = open(path + '/' + root_dirnames[i] + '/' + filename)
                data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                # Import the JSON to a Pandas DF
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[:]
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in names.csv, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
                if not temp_df.empty:
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
            i += 1
    # Replace missing values by 0 (the stat has simply not been initialized because the associated action was not performed)
    df = df.fillna(0)
    if csvtoggle == "true":
        df.to_csv(csvpath)
    return df


def most_pokemons_leaderboard(df, config):
    # Load the Excel file
    file_path = "output.xlsx"
    wb = openpyxl.load_workbook(file_path)
    
    sheet_name = "Global"
    ws = wb[sheet_name]
    i = 0
    ExcelRows = int(config['LEADERBOARD']['ExcelRows'])
    ExcelCols = int(config['LEADERBOARD']['ExcelColumns'])
    for index, row in df[0:ExcelRows*ExcelCols].iterrows():
        ws.cell(row=(i%ExcelRows)+3, column=2+math.floor(i/ExcelRows)*3, value=str(i+1)+".")
        ws.cell(row=(i%ExcelRows)+3, column=3+math.floor(i/ExcelRows)*3, value=index)
        ws.cell(row=(i%ExcelRows)+3, column=4+math.floor(i/ExcelRows)*3, value=row[0])
        i += 1
    now = datetime.datetime.now()
    ws.cell(row=ExcelRows+3, column=2, value="Dernière update le "+now.strftime("%d.%m.%y à %H:%M"))
    ws.cell(row=ExcelRows+4, column=2, value=config['LEADERBOARD']['Subtitle'])
    wb.save(file_path)

def shiny_pokemons_leaderboard(df, config):
    # Load the Excel file
    file_path = "output.xlsx"
    wb = openpyxl.load_workbook(file_path)
    
    sheet_name = "Shiny"
    ws = wb[sheet_name]
    i = 0
    ExcelRows = int(config['SHINYLEADERBOARD']['ExcelRows'])
    ExcelCols = int(config['SHINYLEADERBOARD']['ExcelColumns'])
    for index, row in df[0:ExcelRows*ExcelCols].iterrows():
        ws.cell(row=(i%ExcelRows)+3, column=2+math.floor(i/ExcelRows)*3, value=str(i+1)+".")
        ws.cell(row=(i%ExcelRows)+3, column=3+math.floor(i/ExcelRows)*3, value=index)
        ws.cell(row=(i%ExcelRows)+3, column=4+math.floor(i/ExcelRows)*3, value=row[0])
        i += 1
    now = datetime.datetime.now()
    ws.cell(row=ExcelRows+3, column=2, value="Dernière update le "+now.strftime("%d.%m.%y à %H:%M"))
    ws.cell(row=ExcelRows+4, column=2, value=config['SHINYLEADERBOARD']['Subtitle'])
    wb.save(file_path)


# Read config
config = configparser.ConfigParser()
config.read('cobblemon_config.ini')

# Connect to SFTP if activated
sftp_client = None
if config['FTP']['UseFTP'] == "true":
    transport = paramiko.Transport((config['FTP']['Host'], int(config['FTP']['Port'])))
    transport.connect(username=open("../username.txt", "r").read().strip(), password=open("../password.txt", "r").read().strip())
    sftp_client = paramiko.SFTPClient.from_transport(transport)

# Load the data
if config['GLOBALMATRIX']['UseCSV'] == "false":
    df = loadData(config['GLOBALMATRIX']['CreateCSV'], config['GLOBALMATRIX']['CSVPath'], config['FTP']['UseFTP'], sftp_client, config['FTP']['Path'])
else:
    df = pd.read_csv(config['GLOBALMATRIX']['CSVPath'], index_col=[0,1,2], skipinitialspace=True)

# Close the Connection
if sftp_client:
    sftp_client.close()

# Prepare the counting DF
count_df = df.drop(['caughtTimestamp', 'discoveredTimestamp', 'isShiny'], level=2)
count_df['times_caught'] = count_df.apply(lambda row: (row == "CAUGHT").sum(), axis=1)
#print(count_df['times_caught'].sort_values().to_string())
count_df.drop('times_caught', axis=1, inplace=True)

# Leaderboard feature
if config['LEADERBOARD']['Enable'] == "true":
    player_sum = pd.DataFrame((count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['LEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard(player_sum.iloc, config)

# Shiny leaderboard feature
if config['SHINYLEADERBOARD']['Enable'] == "true":
    player_sum = pd.DataFrame(((df == "True") | (df == True)).sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['SHINYLEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    shiny_pokemons_leaderboard(player_sum.iloc, config)
